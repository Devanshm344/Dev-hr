import csv
import io
from datetime import date, timedelta
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func as sa_func
from typing import Optional, List
from pydantic import BaseModel
from app.db.database import get_db, get_auth_db
from app.models.base import Department, Employee, UserRole, User, UserActivityLog, StatusEnum
from app.core.security import get_current_admin, get_effective_role, set_system_role
from app.api.routes.employees import trend, check_circular_hierarchy, _hard_delete_employee
from app.api.routes.auth import issue_password_reset_token
from app.services.activity_log import log_activity as _log_activity
from app.engagement.core.email import send_email

router = APIRouter()

# Canonical roles in the RBAC system. Super Admin is a superset of Admin
# (passes every Admin-gated check) with a distinct label for the one
# system-owner account.
VALID_RBAC_ROLES = {"Super Admin", "Admin", "Employee"}


def _user_to_dict(emp: Employee, effective_role: str, is_active: bool = True, last_login=None, last_active=None) -> dict:
    return {
        "id":             emp.id,
        "employee_id":    emp.employee_id,
        "name":           f"{emp.first_name} {emp.last_name}",
        "first_name":     emp.first_name,
        "last_name":      emp.last_name,
        "email":          emp.email,
        "role":           effective_role,
        "status":         emp.status.value if emp.status else None,
        "department":     emp.department.name if emp.department else None,
        "title":          emp.title,
        "employment_type": emp.employment_type,
        "work_location":  emp.work_location,
        "manager_id":     emp.manager_id,
        "reporting_manager_name": emp.reporting_manager_name,
        "date_of_joining": str(emp.date_of_joining) if emp.date_of_joining else None,
        "is_active":      is_active,
        "last_login":     last_login.isoformat() if last_login else None,
        "last_active":    last_active.isoformat() if last_active else None,
        "timezone":       emp.timezone or "Asia/Kolkata",
        "country":        emp.country,
        "mfa_enabled":    emp.mfa_enabled or False,
    }


@router.get("/users")
def list_users(
    search: Optional[str] = None,
    role:   Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    department_id: Optional[List[int]] = Query(None),
    title: Optional[List[str]] = Query(None),
    employment_type: Optional[List[str]] = Query(None),
    manager_id: Optional[List[int]] = Query(None),
    work_location: Optional[List[str]] = Query(None),
    joined_after: Optional[date] = None,
    joined_before: Optional[date] = None,
    sort_by: str = "name",
    sort_dir: str = "asc",
    page: int = 1,
    page_size: int = 10,
    db:         Session = Depends(get_db),
    auth_db:    Session = Depends(get_auth_db),
    current_user: Employee = Depends(get_current_admin),
):
    page_size = min(max(page_size, 1), 200)
    page = max(page, 1)

    query = db.query(Employee)
    if search:
        query = query.filter(
            or_(
                Employee.first_name.ilike(f"%{search}%"),
                Employee.last_name.ilike(f"%{search}%"),
                Employee.email.ilike(f"%{search}%"),
                Employee.employee_id.ilike(f"%{search}%"),
                Employee.title.ilike(f"%{search}%"),
            )
        )
    if status:
        query = query.filter(Employee.status.in_(status))
    if department_id:
        query = query.filter(Employee.department_id.in_(department_id))
    if title:
        query = query.filter(Employee.title.in_(title))
    if employment_type:
        query = query.filter(Employee.employment_type.in_(employment_type))
    if manager_id:
        query = query.filter(Employee.manager_id.in_(manager_id))
    if work_location:
        query = query.filter(Employee.work_location.in_(work_location))
    if joined_after:
        query = query.filter(Employee.date_of_joining >= joined_after)
    if joined_before:
        query = query.filter(Employee.date_of_joining <= joined_before)

    employees = query.all()

    # Build effective-role map from user_roles (auth DB)
    role_records = {ur.employee_id: ur.system_role for ur in auth_db.query(UserRole).all()}
    login_map = {
        u.emp_id: (u.is_active, u.last_login, u.last_active)
        for u in db.query(User).all() if u.emp_id is not None
    }

    result = []
    for emp in employees:
        effective_role = role_records.get(emp.employee_id, "Employee")
        if role and effective_role not in role:
            continue
        is_active, last_login, last_active = login_map.get(emp.id, (True, None, None))
        result.append(_user_to_dict(emp, effective_role, is_active, last_login, last_active))

    reverse = sort_dir == "desc"
    sort_key_fns = {
        "name": lambda u: u["name"],
        "department": lambda u: u["department"] or "",
        "role": lambda u: u["role"],
        "status": lambda u: u["status"] or "",
        "joined": lambda u: u["date_of_joining"] or "",
        "employmentType": lambda u: u["employment_type"] or "",
    }
    result.sort(key=sort_key_fns.get(sort_by, sort_key_fns["name"]), reverse=reverse)

    total = len(result)
    start = (page - 1) * page_size
    items = result[start:start + page_size]

    return {
        "items": items,
        "total": total,
        "page": page,
        "pageSize": page_size,
        "totalPages": max(1, (total + page_size - 1) // page_size),
    }


@router.get("/filters")
def get_filters(
    db:      Session = Depends(get_db),
    current_user: Employee = Depends(get_current_admin),
):
    """Real distinct values only — no hardcoded lists — so the UI's filter
    dropdowns only ever offer values that actually exist in dev-hr."""
    departments = db.query(Department).order_by(Department.name).all()
    titles = sorted({t for (t,) in db.query(Employee.title).distinct().all() if t})
    employment_types = sorted({t for (t,) in db.query(Employee.employment_type).distinct().all() if t})
    work_locations = sorted({t for (t,) in db.query(Employee.work_location).distinct().all() if t})
    managers = (
        db.query(Employee)
        .filter(Employee.id.in_(db.query(Employee.manager_id).filter(Employee.manager_id.isnot(None))))
        .order_by(Employee.first_name, Employee.last_name)
        .all()
    )
    return {
        "departments": [{"id": d.id, "name": d.name} for d in departments],
        "designations": titles,
        "employmentTypes": employment_types,
        "workLocations": work_locations,
        "managers": [{"id": m.id, "name": f"{m.first_name} {m.last_name}"} for m in managers],
        "statuses": ["active", "inactive", "on_leave", "terminated"],
        "roles": sorted(VALID_RBAC_ROLES),
    }


def _get_employee_or_404(db: Session, employee_id: int) -> Employee:
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    return emp


def _login_state(db: Session, emp: Employee):
    row = db.query(User).filter(User.emp_id == emp.id).first()
    return (row.is_active, row.last_login, row.last_active) if row else (True, None, None)


def _get_or_create_login_row(db: Session, emp: Employee) -> User:
    row = db.query(User).filter(User.emp_id == emp.id).first()
    if row is None:
        row = User(emp_id=emp.id, employee_id=emp.employee_id, full_name=f"{emp.first_name} {emp.last_name}", email=emp.email, is_active=True)
        db.add(row)
        db.flush()
    return row


@router.get("/users/{employee_id}")
def get_user_detail(
    employee_id: int,
    db:      Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: Employee = Depends(get_current_admin),
):
    emp = _get_employee_or_404(db, employee_id)
    role = get_effective_role(emp, auth_db)
    is_active, last_login, last_active = _login_state(db, emp)
    return _user_to_dict(emp, role, is_active, last_login, last_active)


@router.get("/users/{employee_id}/activity")
def get_user_activity(
    employee_id: int,
    limit: int = 20,
    db:      Session = Depends(get_db),
    current_user: Employee = Depends(get_current_admin),
):
    _get_employee_or_404(db, employee_id)
    rows = (
        db.query(UserActivityLog)
        .filter(UserActivityLog.employee_id == employee_id)
        .order_by(UserActivityLog.created_at.desc())
        .limit(limit)
        .all()
    )
    actor_ids = {r.actor_id for r in rows if r.actor_id}
    actor_names = {}
    if actor_ids:
        for a in db.query(Employee).filter(Employee.id.in_(actor_ids)).all():
            actor_names[a.id] = f"{a.first_name} {a.last_name}"
    return [
        {
            "id": r.id,
            "action": r.action,
            "detail": r.detail,
            "actor": actor_names.get(r.actor_id),
            "createdAt": r.created_at.isoformat() if r.created_at else None,
        }
        for r in rows
    ]


class RoleUpdate(BaseModel):
    role: str


@router.put("/users/{employee_id}/role")
def update_user_role(
    employee_id: int,
    data: RoleUpdate,
    db:         Session = Depends(get_db),
    auth_db:    Session = Depends(get_auth_db),
    current_user: Employee = Depends(get_current_admin),
):
    if data.role not in VALID_RBAC_ROLES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid role. Must be one of: {', '.join(sorted(VALID_RBAC_ROLES))}"
        )

    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    if emp.id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot change your own role")

    set_system_role(auth_db, emp, data.role)
    _log_activity(db, emp.id, current_user.id, "role_change", f"Role changed to {data.role}")
    db.commit()

    return _user_to_dict(emp, data.role)


class AccessUpdate(BaseModel):
    is_active: bool


@router.put("/users/{employee_id}/access")
def update_user_access(
    employee_id: int,
    data: AccessUpdate,
    db:      Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: Employee = Depends(get_current_admin),
):
    emp = _get_employee_or_404(db, employee_id)

    if emp.id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot change your own login access")

    row = _get_or_create_login_row(db, emp)
    row.is_active = data.is_active
    _log_activity(db, emp.id, current_user.id, "login_access_change", "Enabled" if data.is_active else "Disabled")
    db.commit()

    role = get_effective_role(emp, auth_db)
    return _user_to_dict(emp, role, is_active=data.is_active, last_login=row.last_login, last_active=row.last_active)


class ArchiveUpdate(BaseModel):
    archived: bool


@router.put("/users/{employee_id}/archive")
def update_user_archive(
    employee_id: int,
    data: ArchiveUpdate,
    db:      Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: Employee = Depends(get_current_admin),
):
    """Archive is non-destructive: it reuses the existing 'inactive' employee
    status (already a first-class value in filters/KPIs) rather than deleting
    anything, so it's fully reversible via Unarchive."""
    emp = _get_employee_or_404(db, employee_id)

    if emp.id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot archive your own account")

    emp.status = StatusEnum.inactive if data.archived else StatusEnum.active
    _log_activity(db, emp.id, current_user.id, "employee_archived" if data.archived else "employee_unarchived")
    db.commit()

    role = get_effective_role(emp, auth_db)
    is_active, last_login, last_active = _login_state(db, emp)
    return _user_to_dict(emp, role, is_active, last_login, last_active)


@router.put("/users/{employee_id}/mfa")
def admin_disable_mfa(
    employee_id: int,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_admin),
):
    """Recovery path for a user locked out of their authenticator — an admin
    can only DISABLE MFA (never enable it on someone's behalf, since enabling
    requires the user's own authenticator app to complete enrollment)."""
    emp = _get_employee_or_404(db, employee_id)
    emp.mfa_enabled = False
    emp.mfa_secret = None
    _log_activity(db, emp.id, current_user.id, "mfa_disabled_by_admin")
    db.commit()
    return {"mfa_enabled": False}


# ── Bulk actions (User Management's multi-select toolbar) ──────────────────
# Each endpoint below processes one id at a time inside its own try/except so
# a single failure (not found, self-guard, validation) doesn't abort the rest
# of the batch — callers get a per-id result list back instead of an
# all-or-nothing error.

class BulkIdsRequest(BaseModel):
    ids: List[int]
    notify_email: bool = False


class BulkAccessRequest(BaseModel):
    ids: List[int]
    is_active: bool


class BulkRoleRequest(BaseModel):
    ids: List[int]
    role: str


class BulkManagerRequest(BaseModel):
    ids: List[int]
    manager_id: Optional[int] = None


@router.post("/users/bulk/access")
def bulk_update_access(
    data: BulkAccessRequest,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_admin),
):
    results = []
    for eid in data.ids:
        if eid == current_user.id:
            results.append({"employee_id": eid, "success": False, "error": "Cannot change your own login access"})
            continue
        emp = db.query(Employee).filter(Employee.id == eid).first()
        if not emp:
            results.append({"employee_id": eid, "success": False, "error": "Not found"})
            continue
        row = _get_or_create_login_row(db, emp)
        row.is_active = data.is_active
        _log_activity(db, emp.id, current_user.id, "login_access_change", "Enabled" if data.is_active else "Disabled")
        results.append({"employee_id": eid, "success": True})
    db.commit()
    return {"results": results}


@router.post("/users/bulk/role")
def bulk_update_role(
    data: BulkRoleRequest,
    db: Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: Employee = Depends(get_current_admin),
):
    if data.role not in VALID_RBAC_ROLES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid role. Must be one of: {', '.join(sorted(VALID_RBAC_ROLES))}"
        )
    results = []
    for eid in data.ids:
        if eid == current_user.id:
            results.append({"employee_id": eid, "success": False, "error": "Cannot change your own role"})
            continue
        emp = db.query(Employee).filter(Employee.id == eid).first()
        if not emp:
            results.append({"employee_id": eid, "success": False, "error": "Not found"})
            continue
        set_system_role(auth_db, emp, data.role)
        _log_activity(db, emp.id, current_user.id, "role_change", f"Role changed to {data.role}")
        results.append({"employee_id": eid, "success": True})
    db.commit()
    return {"results": results}


@router.post("/users/bulk/manager")
def bulk_assign_manager(
    data: BulkManagerRequest,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_admin),
):
    mgr = None
    if data.manager_id:
        mgr = db.query(Employee).filter(Employee.id == data.manager_id).first()
        if not mgr:
            raise HTTPException(status_code=404, detail="Manager not found")

    results = []
    for eid in data.ids:
        if data.manager_id and eid == data.manager_id:
            results.append({"employee_id": eid, "success": False, "error": "Employee cannot be their own manager"})
            continue
        emp = db.query(Employee).filter(Employee.id == eid).first()
        if not emp:
            results.append({"employee_id": eid, "success": False, "error": "Not found"})
            continue
        if data.manager_id and check_circular_hierarchy(db, eid, data.manager_id):
            results.append({"employee_id": eid, "success": False, "error": "Would create a circular reporting hierarchy"})
            continue
        emp.manager_id = data.manager_id
        emp.reporting_manager_name = f"{mgr.first_name} {mgr.last_name}" if mgr else None
        _log_activity(db, emp.id, current_user.id, "manager_change", f"Manager set to {mgr.first_name + ' ' + mgr.last_name if mgr else 'None'}")
        results.append({"employee_id": eid, "success": True})
    db.commit()
    return {"results": results}


@router.post("/users/bulk/reset-password")
def bulk_reset_password(
    data: BulkIdsRequest,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_admin),
):
    """Used by both the toolbar's Reset Password and Invite buttons — same
    underlying token-issue mechanism. The link is always shown to the admin as a
    fallback/confirmation; when notify_email is set (the Invite flow) it's also
    emailed directly to the employee via send_email(), which logs instead of
    sending on a checkout with no SMTP_HOST configured (see engagement/core/email.py)."""
    results = []
    for eid in data.ids:
        emp = db.query(Employee).filter(Employee.id == eid).first()
        if not emp:
            results.append({"employee_id": eid, "success": False, "error": "Not found"})
            continue
        reset_link = issue_password_reset_token(db, emp)
        _log_activity(db, emp.id, current_user.id, "reset_password_link_generated")
        result = {"employee_id": eid, "name": f"{emp.first_name} {emp.last_name}", "success": True, "reset_link": reset_link}
        if data.notify_email and emp.email:
            emailed = send_email(
                emp.email,
                "You're invited to Cotelligent HRMS",
                f"Hi {emp.first_name},\n\nAn account has been set up for you on Cotelligent HRMS. "
                f"Use the link below to set your password and sign in:\n\n{reset_link}\n\n"
                f"This link expires shortly — if it's expired, ask an admin to resend your invitation.",
            )
            result["emailed"] = emailed
            _log_activity(db, emp.id, current_user.id, "invitation_emailed" if emailed else "invitation_email_failed")
        results.append(result)
    db.commit()
    return {"results": results}


@router.post("/users/bulk/delete")
def bulk_delete_users(
    data: BulkIdsRequest,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_admin),
):
    """Irreversible — wraps the same cascade-delete used by employees.py's
    single-employee endpoint (_hard_delete_employee), one id at a time so a
    failure partway through doesn't silently roll back earlier successes."""
    results = []
    for eid in data.ids:
        if eid == current_user.id:
            results.append({"employee_id": eid, "success": False, "error": "Cannot delete your own account"})
            continue
        emp = db.query(Employee).filter(Employee.id == eid).first()
        name = f"{emp.first_name} {emp.last_name}" if emp else None
        try:
            _hard_delete_employee(db, eid)
            results.append({"employee_id": eid, "name": name, "success": True})
        except HTTPException as e:
            results.append({"employee_id": eid, "name": name, "success": False, "error": e.detail})
    return {"results": results}


@router.get("/kpis")
def get_user_kpis(
    db:      Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: Employee = Depends(get_current_admin),
):
    """KPI-card row for the User Management page. Every figure below is a
    live count/derivation from employees/users/user_roles — no placeholder
    or fabricated data. Snapshot-only fields (active/inactive/locked/admin)
    have no trend/sparkline, since this schema keeps no daily history of
    status/role/lock changes to chart — only date_of_joining-based figures
    (total, new-this-month) have a genuine trailing window to show.
    """
    today = date.today()
    days = [today - timedelta(days=i) for i in range(6, -1, -1)]  # last 7 days, oldest -> today
    month_start = today.replace(day=1)
    prev_month_end = month_start - timedelta(days=1)
    prev_month_start = prev_month_end.replace(day=1)

    total_users = db.query(Employee).count()
    active_users = db.query(Employee).filter(Employee.status == "active").count()
    inactive_users = db.query(Employee).filter(Employee.status.in_(["inactive", "on_leave"])).count()

    locked_accounts = (
        db.query(User)
        .join(Employee, Employee.id == User.emp_id)
        .filter(User.is_active.is_(False))
        .count()
    )

    admin_users = (
        auth_db.query(UserRole)
        .filter(UserRole.system_role.in_(["Admin", "Super Admin"]))
        .count()
    )

    # -- Growth trend: same date_of_joining grouping employees/stats uses --
    hire_rows = (
        db.query(Employee.date_of_joining, sa_func.count(Employee.id))
        .filter(Employee.date_of_joining >= days[0], Employee.date_of_joining <= today)
        .group_by(Employee.date_of_joining)
        .all()
    )
    hire_counts = dict(hire_rows)
    running_total = db.query(Employee).filter(Employee.date_of_joining < days[0]).count()
    total_sparkline, new_hire_sparkline = [], []
    for d in days:
        c = hire_counts.get(d, 0)
        running_total += c
        total_sparkline.append(running_total)
        new_hire_sparkline.append(c)

    new_this_month = db.query(Employee).filter(
        Employee.date_of_joining >= month_start, Employee.date_of_joining <= today
    ).count()
    new_last_month = db.query(Employee).filter(
        Employee.date_of_joining >= prev_month_start, Employee.date_of_joining <= prev_month_end
    ).count()

    return {
        "totalUsers":      {"value": total_users,      "trend": trend(new_this_month, "this month", total_sparkline)},
        "activeUsers":     {"value": active_users},
        "inactiveUsers":   {"value": inactive_users},
        "lockedAccounts":  {"value": locked_accounts},
        "adminUsers":      {"value": admin_users},
        "newThisMonth":    {"value": new_this_month, "trend": trend(new_this_month - new_last_month, "vs last month", new_hire_sparkline)},
    }


@router.get("/charts")
def get_user_charts(
    db:      Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: Employee = Depends(get_current_admin),
):
    """Analytics for the User Management page — same real-data-only rule
    as /kpis. Department-distribution query mirrors departments.py's own
    get_department_distribution() shape rather than inventing a new one.
    """
    dept_rows = (
        db.query(Department, sa_func.count(Employee.id))
        .outerjoin(Employee, Employee.department_id == Department.id)
        .group_by(Department.id)
        .all()
    )
    total = sum(count for _, count in dept_rows)
    by_department = sorted(
        [{"name": d.name, "value": count} for d, count in dept_rows],
        key=lambda x: x["value"], reverse=True,
    )

    role_counts: dict = {}
    all_roles = {ur.employee_id: ur.system_role for ur in auth_db.query(UserRole).all()}
    for emp_id, _ in db.query(Employee.employee_id, Employee.id).all():
        r = all_roles.get(emp_id, "Employee")
        role_counts[r] = role_counts.get(r, 0) + 1
    by_role = sorted(
        [{"name": k, "value": v} for k, v in role_counts.items()],
        key=lambda x: x["value"], reverse=True,
    )

    status_rows = (
        db.query(Employee.status, sa_func.count(Employee.id))
        .group_by(Employee.status)
        .all()
    )
    active_vs_inactive = [
        {"name": (s.value if s else "unknown"), "value": c} for s, c in status_rows
    ]

    month_rows = (
        db.query(sa_func.to_char(Employee.date_of_joining, 'YYYY-MM'), sa_func.count(Employee.id))
        .filter(Employee.date_of_joining.isnot(None))
        .group_by(sa_func.to_char(Employee.date_of_joining, 'YYYY-MM'))
        .order_by(sa_func.to_char(Employee.date_of_joining, 'YYYY-MM'))
        .all()
    )
    monthly_additions = [{"month": m, "value": v} for m, v in month_rows[-12:]]

    login_day_rows = (
        db.query(sa_func.to_char(UserActivityLog.created_at, 'YYYY-MM-DD'), sa_func.count(UserActivityLog.id))
        .filter(UserActivityLog.action == "login")
        .group_by(sa_func.to_char(UserActivityLog.created_at, 'YYYY-MM-DD'))
        .order_by(sa_func.to_char(UserActivityLog.created_at, 'YYYY-MM-DD'))
        .all()
    )
    login_activity = [{"day": d, "value": v} for d, v in login_day_rows[-14:]]

    return {
        "usersByDepartment": by_department,
        "usersByRole": by_role,
        "activeVsInactive": active_vs_inactive,
        "monthlyAdditions": monthly_additions,
        "loginActivity": login_activity,
        "departmentDistribution": [
            {"name": r["name"], "value": round(r["value"] / total * 100, 1) if total else 0.0}
            for r in by_department
        ],
    }


@router.get("/recently-added")
def get_recently_added(
    limit: int = 6,
    db:      Session = Depends(get_db),
    current_user: Employee = Depends(get_current_admin),
):
    employees = (
        db.query(Employee)
        .filter(Employee.date_of_joining.isnot(None))
        .order_by(Employee.date_of_joining.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "id": e.id,
            "name": f"{e.first_name} {e.last_name}",
            "department": e.department.name if e.department else None,
            "title": e.title,
            "dateOfJoining": str(e.date_of_joining),
            "profilePicture": e.profile_picture,
        }
        for e in employees
    ]


EXPORT_COLUMNS = [
    ("employee_id", "Employee ID"), ("name", "Name"), ("email", "Work Email"),
    ("department", "Department"), ("title", "Designation"), ("role", "Role"),
    ("employment_type", "Employment Type"), ("status", "Status"),
    ("work_location", "Office Location"), ("date_of_joining", "Joining Date"),
    ("reporting_manager_name", "Reporting Manager"),
]

_FORMULA_LEAD_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _csv_safe(value):
    """Neutralizes CSV/Excel formula injection: a leading =, +, -, @ (or tab/CR)
    makes spreadsheet apps evaluate the cell as a formula on open. Prefixing
    with a single quote forces it to be read as plain text instead."""
    if isinstance(value, str) and value.startswith(_FORMULA_LEAD_CHARS):
        return "'" + value
    return value


def _export_rows(
    db: Session, auth_db: Session,
    department_id: Optional[List[int]], role: Optional[List[str]], status: Optional[List[str]],
    work_location: Optional[List[str]], joined_after: Optional[date], joined_before: Optional[date],
    ids: Optional[List[int]] = None,
) -> List[dict]:
    """Same narrower filter set the source project's export endpoints used
    (no search/designation/employmentType/manager) — built on the same
    _user_to_dict shape list_users already uses. `ids` lets the bulk-action
    toolbar's Export button export just the selected rows instead of the
    page's active filters."""
    query = db.query(Employee)
    if ids:
        query = query.filter(Employee.id.in_(ids))
    if department_id:
        query = query.filter(Employee.department_id.in_(department_id))
    if status:
        query = query.filter(Employee.status.in_(status))
    if work_location:
        query = query.filter(Employee.work_location.in_(work_location))
    if joined_after:
        query = query.filter(Employee.date_of_joining >= joined_after)
    if joined_before:
        query = query.filter(Employee.date_of_joining <= joined_before)

    role_records = {ur.employee_id: ur.system_role for ur in auth_db.query(UserRole).all()}
    rows = []
    for emp in query.all():
        emp_role = role_records.get(emp.employee_id, "Employee")
        if role and emp_role not in role:
            continue
        rows.append(_user_to_dict(emp, emp_role))
    return rows


@router.get("/export/csv")
def export_users_csv(
    department_id: Optional[List[int]] = Query(None),
    role: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    work_location: Optional[List[str]] = Query(None),
    joined_after: Optional[date] = None,
    joined_before: Optional[date] = None,
    ids: Optional[List[int]] = Query(None),
    db:      Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: Employee = Depends(get_current_admin),
):
    rows = _export_rows(db, auth_db, department_id, role, status, work_location, joined_after, joined_before, ids)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for _, label in EXPORT_COLUMNS])
    for r in rows:
        writer.writerow([_csv_safe(r.get(key, "")) for key, _ in EXPORT_COLUMNS])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=users_export.csv"},
    )


@router.get("/export/excel")
def export_users_excel(
    department_id: Optional[List[int]] = Query(None),
    role: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    work_location: Optional[List[str]] = Query(None),
    joined_after: Optional[date] = None,
    joined_before: Optional[date] = None,
    ids: Optional[List[int]] = Query(None),
    db:      Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: Employee = Depends(get_current_admin),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    rows = _export_rows(db, auth_db, department_id, role, status, work_location, joined_after, joined_before, ids)
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, (_, label) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, r in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_csv_safe(r.get(key, "")))
    for col_idx in range(1, len(EXPORT_COLUMNS) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_export.xlsx"},
    )


@router.get("/export/pdf")
def export_users_pdf(
    department_id: Optional[List[int]] = Query(None),
    role: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    work_location: Optional[List[str]] = Query(None),
    joined_after: Optional[date] = None,
    joined_before: Optional[date] = None,
    ids: Optional[List[int]] = Query(None),
    db:      Session = Depends(get_db),
    auth_db: Session = Depends(get_auth_db),
    current_user: Employee = Depends(get_current_admin),
):
    from fpdf import FPDF

    rows = _export_rows(db, auth_db, department_id, role, status, work_location, joined_after, joined_before, ids)
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "User Management Export", ln=1)
    col_width = 260 / len(EXPORT_COLUMNS)

    pdf.set_font("Helvetica", "B", 8)
    for _, label in EXPORT_COLUMNS:
        pdf.cell(col_width, 8, label, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    for r in rows:
        for key, _ in EXPORT_COLUMNS:
            pdf.cell(col_width, 7, str(r.get(key, ""))[:28], border=1)
        pdf.ln()

    out = pdf.output(dest="S")
    if isinstance(out, str):
        out = out.encode("latin-1")
    return StreamingResponse(
        io.BytesIO(out), media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=users_export.pdf"},
    )
